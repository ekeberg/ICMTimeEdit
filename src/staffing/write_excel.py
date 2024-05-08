from collections import defaultdict
import pandas
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelWriter:
    """A class used for writing personal dataframes to a larger excel file"""
    def __init__(self, file_name) -> None:
        self.writer = pandas.ExcelWriter(file_name, engine='openpyxl')
        self.startrow = defaultdict(int)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.writer.close()

    def save(self):
        """Save the excel file"""
        self.writer.close()

    def write_person_frame(self, person, frame):
        """Write the personal frame to the excel file"""
        program = person.program
        # Write name and employment
        pandas.DataFrame([person.name, person.employment]).T.to_excel(
            self.writer, sheet_name=program,
            startrow=self.startrow[program],
            index=False, header=False)
        self.startrow[program] += 1

        # Write the dataframe
        frame.to_excel(self.writer, sheet_name=program,
                       startrow=self.startrow[program],
                       index=True, header=True)
        self.startrow[program] += len(frame) + 2

    def write_person_list(self, person, frame):
        """Write the personal frame to the excel file"""
        # Write name and employment
        pandas.DataFrame([person.name, person.employment]).T.to_excel(
            self.writer, sheet_name=person.name,
            startrow=0,
            index=False, header=False)

        # Write the dataframe
        frame.to_excel(self.writer, sheet_name=person.name,
                       startrow=2,
                       index=True, header=True)


def make_columns_larger(filename: str) -> None:
    """Make the columns in the excel file larger"""
    wb = load_workbook(filename)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for col in sheet.columns:
            length = max(len(str(cell.value)) for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = length + 2

    wb.save(filename)
