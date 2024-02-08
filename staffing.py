"""Tools used to analyze the teaching hours for ICM"""
from datetime import datetime
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# How to categorize the different types of events
# EVENT_TYPES = {"Föreläsning": ["Föreläsning", "Introduktion", "Upprop"],
#                "Laboration": ["Laboration", "Datorlab", "Tutorials", "Lablektion",],
#                "Seminarium": ["Seminarium", "Frågestund", "Presentation", "Fall",
#                               "Lektion", "Övning", "Workshop", "Diskussion",
#                               "Tentamen", "Examination", "Dugga", "Case",
#                               "Räkneövning", "Grupparbete", "Symposium", "Projekt",
#                               "Möte", "Problemlösning", "Övrigt", "Redovisning",
#                               "Konferens"],
#                               "Exkursion": ["Studiebesök"]}
EVENT_TYPES = {
    "Föreläsning": ["Föreläsning", "Introduktion", "Upprop"],
    "Laboration": ["Laboration", "Datorlab", "Tutorials"],
    "Exkursion": ["Studiebesök"],
    "Seminarium": ["Seminarium", "Frågestund", "Presentation",
                   "Fall", "Lektion", "Övning", "Workshop",
                   "Diskussion", "Case", "Räkneövning",
                   "Grupparbete", "Symposium", "Projekt",
                   "Problemlösning", "Redovisning"],
}


WEIGHT_FACTORS = {"Föreläsning": 4,
                  "Laboration": 2,
                  "Seminarium": 1.5,
                  "Exkursion": 1.5,
                  "Okänd": 2}
WEIGHT_FACTORS_PHD = WEIGHT_FACTORS.copy()
WEIGHT_FACTORS_PHD["Föreläsning"] = 8



@dataclass(frozen=True)
class Person:
    """A person at ICM"""
    first_name: str
    last_name: str
    employment: str
    program: str

    def __str__(self):
        return f"{self.first_name} {self.last_name}"

    def __repr__(self):
        return f"P({self.first_name} {self.last_name})"

    @property
    def name(self):
        """Return the full name of the person"""
        return f"{self.first_name} {self.last_name}"

def is_phd_student(person: Person) -> bool:
    """Check if a person is a PhD student"""
    return "doktorand" in person.employment.lower()


def time_diff(start: str, end: str) -> int:
    """Calculate the time difference between two times in the format hh:mm"""
    start = datetime.strptime(start, "%H:%M")
    end = datetime.strptime(end, "%H:%M")
    return (end - start).seconds // 60


def minutes_to_lecture_hours(minutes: int) -> float:
    """Convert minutes to lecture hours (45 minuts per lecture hour)"""
    nbreaks = (minutes + 15) // 60 - 1
    adjusted_minutes = minutes - 15*nbreaks
    nlectures = adjusted_minutes / 45
    return nlectures


def name_to_person(name: str, people: list) -> Person:
    """Find the person object from a name"""
    for person in people:
        if name == person.name:
            return person
    raise ValueError(f"Could not find person {name}")


def make_columns_larger(filename: str) -> None:
    """Make the columns in the excel file larger"""
    wb = load_workbook(filename)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for col in sheet.columns:
            length = max(len(str(cell.value)) for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = length + 2

    wb.save(filename)


def get_factors_for_person(person: Person) -> dict:
    """Return the weight factors for a person"""
    if is_phd_student(person):
        return WEIGHT_FACTORS_PHD
    else:
        return WEIGHT_FACTORS


def identify_event_type(event: dict) -> str:
    """Identify the type of event"""
    for event_type, keywords in EVENT_TYPES.items():
        for keyword in keywords:
            if keyword in event["Kursnamn"]:
                return event_type
    return "Okänd"
